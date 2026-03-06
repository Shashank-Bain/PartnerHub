import os
import json
import unicodedata
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path

from flask import Flask, jsonify, request, session
from flask_cors import CORS
from dotenv import load_dotenv
from werkzeug.security import check_password_hash, generate_password_hash

load_dotenv()

app = Flask(__name__)

APP_VERSION = os.getenv("APP_VERSION") or os.getenv("VERCEL_GIT_COMMIT_SHA") or "dev"

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")

_configured_data_dir = os.getenv("PARTNERHUB_DATA_DIR", "").strip()
if _configured_data_dir:
    USERS_DATA_PATH = os.path.join(_configured_data_dir, "users.json")
elif os.path.isdir(DATA_DIR) and os.access(DATA_DIR, os.W_OK):
    USERS_DATA_PATH = os.path.join(DATA_DIR, "users.json")
else:
    USERS_DATA_PATH = os.path.join("/tmp", "partnerhub", "users.json")

app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "pulse-partner-hub-dev-secret")

CORS(app, supports_credentials=True, origins=["http://localhost:5173", "http://127.0.0.1:5173"])


SECTOR_COMPANY_MAP = {
    "Consumer Products": ["Nestlé", "Unilever", "PepsiCo", "Kraft Heinz", "Danone"],
    "Mining": ["Anglo American", "BHP", "Rio Tinto", "Teck", "Glencore"],
    "Construction Materials": ["Heidelberg Materials", "Holcim", "Cemex", "CRH"],
    "Private Equity": ["CVC", "EQT", "Blackstone", "KKR"],
}

COMPANY_ALIASES = {
    "pepsico": "pepsi",
    "kraftheinz": "kraftheinz",
    "kraft-heinz": "kraftheinz",
    "nestle": "nestle",
    "nestlé": "nestle",
}

BUCKET_BASE = {
    "X": 0,
    "A": 20,
    "P": 40,
    "L": 60,
    "D": 80,
}

BUCKET_LABELS = {
    "X": "No Commitments",
    "A": "Active",
    "P": "Proactive",
    "L": "Leading",
    "D": "Distinctive",
}


class UserRecord:
    def __init__(self, payload):
        self.id = int(payload.get("id"))
        self.email = str(payload.get("email") or "").strip().lower()
        self.password_hash = str(payload.get("passwordHash") or "")
        self.first_name = str(payload.get("firstName") or "").strip()
        self.last_name = str(payload.get("lastName") or "").strip()
        self.must_change_password = False
        self.is_admin = bool(payload.get("isAdmin", False))
        self.created_at = str(payload.get("createdAt") or datetime.now(timezone.utc).isoformat())
        self.sectors = sorted({
            str(name or "").strip()
            for name in (payload.get("sectors") or [])
            if str(name or "").strip() in SECTOR_COMPANY_MAP
        })

    def to_dict(self):
        return {
            "id": self.id,
            "email": self.email,
            "firstName": self.first_name,
            "lastName": self.last_name,
            "mustChangePassword": self.must_change_password,
            "isAdmin": self.is_admin,
            "sectors": self.sectors,
        }

    def to_admin_dict(self):
        payload = self.to_dict()
        payload["maskedPassword"] = "********"
        return payload


def _safe_read_json(path, default_value):
    if not os.path.exists(path):
        return default_value
    try:
        with open(path, "r", encoding="utf-8") as data_file:
            return json.load(data_file)
    except (OSError, json.JSONDecodeError):
        return default_value


def _atomic_write_json(path, payload):
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temp_path = target.with_suffix(f"{target.suffix}.tmp")
    with open(temp_path, "w", encoding="utf-8") as data_file:
        json.dump(payload, data_file, ensure_ascii=False, indent=2)
    os.replace(temp_path, target)


def _default_admin_record(user_id=1):
    return {
        "id": int(user_id),
        "email": "admin@pulseph.com",
        "passwordHash": generate_password_hash("admin123"),
        "firstName": "Admin",
        "lastName": "User",
        "mustChangePassword": False,
        "isAdmin": True,
        "sectors": list(SECTOR_COMPANY_MAP.keys()),
        "createdAt": datetime.now(timezone.utc).isoformat(),
    }


def _normalize_user_payload(raw_user):
    return UserRecord(raw_user).to_dict() | {
        "passwordHash": str(raw_user.get("passwordHash") or ""),
        "createdAt": str(raw_user.get("createdAt") or datetime.now(timezone.utc).isoformat()),
    }


def ensure_user_store():
    store = _safe_read_json(USERS_DATA_PATH, {"nextUserId": 1, "users": []})
    needs_write = not os.path.exists(USERS_DATA_PATH)

    if not isinstance(store, dict):
        store = {"nextUserId": 1, "users": []}
        needs_write = True

    users = store.get("users")
    if not isinstance(users, list):
        users = []
        needs_write = True

    normalized_users = []
    seen_emails = set()
    max_user_id = 0

    for raw_user in users:
        if not isinstance(raw_user, dict):
            needs_write = True
            continue
        normalized = _normalize_user_payload(raw_user)
        if normalized != raw_user:
            needs_write = True
        email = normalized["email"]
        if not email or email in seen_emails:
            needs_write = True
            continue
        seen_emails.add(email)
        normalized_users.append(normalized)
        max_user_id = max(max_user_id, int(normalized["id"]))

    admin_user = next((user for user in normalized_users if user.get("email") == "admin@pulseph.com"), None)
    if not admin_user:
        admin_payload = _default_admin_record(max_user_id + 1 if max_user_id else 1)
        normalized_users.append(admin_payload)
        max_user_id = max(max_user_id, admin_payload["id"])
        needs_write = True
    else:
        original_is_admin = bool(admin_user.get("isAdmin", False))
        original_sectors = list(admin_user.get("sectors") or [])
        admin_user["isAdmin"] = True
        admin_user["sectors"] = list(SECTOR_COMPANY_MAP.keys())
        if not original_is_admin or original_sectors != admin_user["sectors"]:
            needs_write = True

    next_user_id = int(store.get("nextUserId") or (max_user_id + 1))
    if next_user_id <= max_user_id:
        next_user_id = max_user_id + 1
        needs_write = True

    normalized_store = {
        "nextUserId": next_user_id,
        "users": sorted(normalized_users, key=lambda user: user["email"]),
    }
    if needs_write:
        _atomic_write_json(USERS_DATA_PATH, normalized_store)
    return normalized_store


def read_user_store():
    store = _safe_read_json(USERS_DATA_PATH, None)
    if not isinstance(store, dict):
        return ensure_user_store()
    if not isinstance(store.get("users"), list):
        return ensure_user_store()
    return store


def write_user_store(store):
    _atomic_write_json(USERS_DATA_PATH, store)


def is_ephemeral_user_store():
    normalized_path = USERS_DATA_PATH.replace("\\", "/").lower()
    return normalized_path.startswith("/tmp/")


def list_users():
    store = read_user_store()
    return [UserRecord(user) for user in store.get("users", [])]


def get_user_by_id(user_id):
    for user in list_users():
        if user.id == int(user_id):
            return user
    return None


def get_user_by_email(email):
    lookup_email = str(email or "").strip().lower()
    if not lookup_email:
        return None
    for user in list_users():
        if user.email == lookup_email:
            return user
    return None


def save_user(user_record):
    store = read_user_store()
    users = store.get("users", [])

    updated_payload = {
        "id": user_record.id,
        "email": user_record.email,
        "passwordHash": user_record.password_hash,
        "firstName": user_record.first_name,
        "lastName": user_record.last_name,
        "mustChangePassword": user_record.must_change_password,
        "isAdmin": user_record.is_admin,
        "sectors": [sector for sector in user_record.sectors if sector in SECTOR_COMPANY_MAP],
        "createdAt": user_record.created_at,
    }

    replaced = False
    for index, existing in enumerate(users):
        if int(existing.get("id", -1)) == user_record.id:
            users[index] = updated_payload
            replaced = True
            break

    if not replaced:
        users.append(updated_payload)

    store["users"] = sorted(users, key=lambda user: str(user.get("email") or ""))
    next_id = int(store.get("nextUserId") or 1)
    if user_record.id >= next_id:
        store["nextUserId"] = user_record.id + 1
    write_user_store(store)


def delete_user(user_id):
    store = read_user_store()
    current_users = store.get("users", [])
    filtered_users = [user for user in current_users if int(user.get("id", -1)) != int(user_id)]
    if len(filtered_users) == len(current_users):
        return False
    store["users"] = filtered_users
    write_user_store(store)
    return True


def create_user_record(email, password, first_name, last_name, sector_names, is_admin=False):
    store = read_user_store()
    users = store.get("users", [])
    normalized_email = email.strip().lower()

    if any(str(user.get("email") or "").strip().lower() == normalized_email for user in users):
        return None

    next_id = int(store.get("nextUserId") or 1)
    user_record = UserRecord(
        {
            "id": next_id,
            "email": normalized_email,
            "passwordHash": generate_password_hash(password),
            "firstName": first_name.strip(),
            "lastName": last_name.strip(),
            "mustChangePassword": False,
            "isAdmin": bool(is_admin),
            "sectors": [sector for sector in sector_names if sector in SECTOR_COMPANY_MAP],
            "createdAt": datetime.now(timezone.utc).isoformat(),
        }
    )
    save_user(user_record)
    return user_record


def get_current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    return get_user_by_id(user_id)


def get_admin_user_or_error():
    user = get_current_user()
    if not user:
        return None, (jsonify({"message": "Unauthorized."}), 401)
    if not user.is_admin:
        return None, (jsonify({"message": "Admin access required."}), 403)
    return user, None


def resolve_sector_names(sector_names):
    clean_sector_names = sorted({(name or "").strip() for name in (sector_names or []) if (name or "").strip()})
    if not clean_sector_names:
        return [], []

    found_names = [name for name in clean_sector_names if name in SECTOR_COMPANY_MAP]
    missing = [name for name in clean_sector_names if name not in SECTOR_COMPANY_MAP]
    return found_names, missing


def get_user_available_sectors(user):
    sector_order = {sector_name: index for index, sector_name in enumerate(SECTOR_COMPANY_MAP.keys())}

    if user.is_admin:
        sectors = list(SECTOR_COMPANY_MAP.keys())
    else:
        sectors = [sector for sector in (user.sectors or []) if sector in SECTOR_COMPANY_MAP]

    return sorted(sectors, key=lambda sector: (sector_order.get(sector, 999), sector.lower()))


def get_ordered_company_names(sector_name):
    configured_order = {company_name: index for index, company_name in enumerate(SECTOR_COMPANY_MAP.get(sector_name, []))}
    company_names = list(SECTOR_COMPANY_MAP.get(sector_name, []))
    return sorted(company_names, key=lambda company_name: (configured_order.get(company_name, 999), company_name.lower()))


def load_material_topics_dataset():
    data_path = os.path.join(BASE_DIR, "data", "material_topics.json")
    if not os.path.exists(data_path):
        return []

    with open(data_path, "r", encoding="utf-8") as data_file:
        parsed = json.load(data_file)

    if isinstance(parsed, list):
        return parsed
    return []


def load_recommendations_dataset():
    data_path = os.path.join(BASE_DIR, "data", "recommendations.json")
    if not os.path.exists(data_path):
        return []

    with open(data_path, "r", encoding="utf-8") as data_file:
        parsed = json.load(data_file)

    if isinstance(parsed, list):
        return parsed
    return []


def load_commitment_breakdown_dataset():
    data_path = os.path.join(BASE_DIR, "data", "commitment_breakdown.json")
    if not os.path.exists(data_path):
        return []

    with open(data_path, "r", encoding="utf-8") as data_file:
        parsed = json.load(data_file)

    if isinstance(parsed, list):
        return parsed
    return []


def load_scorecard_payload():
    data_path = os.path.join(BASE_DIR, "data", "scorecard.json")
    if not os.path.exists(data_path):
        return {}

    with open(data_path, "r", encoding="utf-8") as data_file:
        parsed = json.load(data_file)

    if isinstance(parsed, dict):
        return parsed
    return {}


def load_scorecard_dataset():
    payload = load_scorecard_payload()
    scores_dataset = payload.get("scores") if isinstance(payload, dict) else None
    if isinstance(scores_dataset, dict):
        return scores_dataset
    if isinstance(payload, dict):
        return payload
    return {}


def load_scorecard_best_practices_dataset():
    payload = load_scorecard_payload()
    best_practices = payload.get("best_practices") if isinstance(payload, dict) else None
    if isinstance(best_practices, list):
        return best_practices
    return []


@lru_cache(maxsize=1)
def load_commitment_scale_dataset():
    data_path = os.path.join(BASE_DIR, "data", "Commitment_Scale.xlsx")
    if not os.path.exists(data_path):
        return {
            "bucketLegend": [{"code": key, "label": value} for key, value in BUCKET_LABELS.items()],
            "industries": {},
        }

    try:
        from openpyxl import load_workbook
    except Exception:
        return {
            "bucketLegend": [{"code": key, "label": value} for key, value in BUCKET_LABELS.items()],
            "industries": {},
        }

    workbook = load_workbook(data_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = [str(cell or "").strip() for cell in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))]
    bucket_columns = {
        "X": 4,
        "A": 5,
        "P": 6,
        "L": 7,
        "D": 8,
    }

    bucket_legend = []
    for code, index in bucket_columns.items():
        header_value = header_row[index] if index < len(header_row) else ""
        label = BUCKET_LABELS[code]
        if "-" in header_value:
            _, parsed_label = header_value.split("-", 1)
            parsed_label = parsed_label.strip()
            if parsed_label:
                label = parsed_label
        bucket_legend.append({"code": code, "label": label})

    industries = {}
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        industry_name = str(row[1] or "").strip()
        theme_name = str(row[2] or "").strip()
        parameter_name = str(row[3] or "").strip()
        if not (industry_name and theme_name and parameter_name):
            continue

        parameter_scale = {
            "parameter": parameter_name,
            "buckets": {
                code: str(row[column_index] or "").strip()
                for code, column_index in bucket_columns.items()
            },
        }

        industry_key = normalize_text_token(industry_name)
        theme_key = normalize_text_token(theme_name)
        industries.setdefault(industry_key, {
            "name": industry_name,
            "themes": {},
        })
        industries[industry_key]["themes"].setdefault(theme_key, {
            "name": theme_name,
            "parameters": [],
        })
        industries[industry_key]["themes"][theme_key]["parameters"].append(parameter_scale)

    return {
        "bucketLegend": bucket_legend,
        "industries": industries,
    }


def get_theme_scale(industry_name, theme_name):
    scale_dataset = load_commitment_scale_dataset()
    industry = scale_dataset.get("industries", {}).get(normalize_text_token(industry_name), {})
    theme = (industry.get("themes") or {}).get(normalize_text_token(theme_name), {})
    return theme.get("parameters", []) if isinstance(theme, dict) else []


def get_industry_scale_example(industry_name):
    scale_dataset = load_commitment_scale_dataset()
    industry = scale_dataset.get("industries", {}).get(normalize_text_token(industry_name), {})
    themes = industry.get("themes") if isinstance(industry, dict) else {}
    if not isinstance(themes, dict):
        return None

    for theme in themes.values():
        parameters = theme.get("parameters") if isinstance(theme, dict) else []
        if not parameters:
            continue
        example_parameter = parameters[0]
        return {
            "industry": industry.get("name") or industry_name,
            "theme": theme.get("name") or "",
            "parameter": example_parameter.get("parameter") or "",
            "buckets": example_parameter.get("buckets") or {},
        }
    return None


def load_investment_deals_dataset():
    data_path = os.path.join(BASE_DIR, "data", "Investment_Deals.json")
    if not os.path.exists(data_path):
        return []

    with open(data_path, "r", encoding="utf-8") as data_file:
        parsed = json.load(data_file)

    if isinstance(parsed, list):
        return parsed
    return []


def load_green_capex_dataset():
    data_path = os.path.join(BASE_DIR, "data", "Green_Capex_Deals.json")
    if not os.path.exists(data_path):
        return []

    with open(data_path, "r", encoding="utf-8") as data_file:
        parsed = json.load(data_file)

    if isinstance(parsed, list):
        return parsed
    return []


def load_investment_rationals_dataset():
    data_path = os.path.join(BASE_DIR, "data", "Investment_Deal_Rationals.json")
    if not os.path.exists(data_path):
        return []

    with open(data_path, "r", encoding="utf-8") as data_file:
        parsed = json.load(data_file)

    if isinstance(parsed, list):
        return parsed
    return []


def normalize_company_name(company_name):
    text_value = unicodedata.normalize("NFKD", (company_name or ""))
    ascii_value = "".join(character for character in text_value if not unicodedata.combining(character))
    normalized = "".join(character.lower() for character in ascii_value if character.isalnum())
    return COMPANY_ALIASES.get(normalized, normalized)


def normalize_text_token(value):
    text_value = unicodedata.normalize("NFKD", (value or ""))
    ascii_value = "".join(character for character in text_value if not unicodedata.combining(character))
    return "".join(character.lower() for character in ascii_value if character.isalnum())


def company_text_matches(candidate_value, selected_company):
    candidate_token = normalize_text_token(candidate_value)
    selected_token = normalize_text_token(selected_company)
    if not candidate_token or not selected_token:
        return False

    if candidate_token == selected_token:
        return True

    if candidate_token.startswith(selected_token):
        return True

    return selected_token in candidate_token


def parse_numeric_value(raw_value):
    text_value = str(raw_value or "").strip()
    if not text_value:
        return None

    normalized = text_value.lower()
    if normalized in {"-", "na", "n/a", "not disclosed", "undisclosed"}:
        return None

    filtered = "".join(character for character in text_value if character.isdigit() or character in {".", "-"})
    if not filtered:
        return None

    try:
        return float(filtered)
    except ValueError:
        return None


def parse_flexible_date(raw_value):
    text_value = str(raw_value or "").strip()
    if not text_value or text_value.lower() in {"na", "n/a", "-"}:
        return None

    supported_formats = [
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%d",
        "%B %Y",
        "%Y",
    ]

    for date_format in supported_formats:
        try:
            return datetime.strptime(text_value, date_format)
        except ValueError:
            continue

    return None


def parse_year(raw_value):
    parsed = parse_flexible_date(raw_value)
    if parsed:
        return parsed.year

    text_value = str(raw_value or "").strip()
    if not text_value:
        return None

    digits = "".join(character for character in text_value if character.isdigit())
    if len(digits) >= 4:
        try:
            year = int(digits[:4])
            if 1900 <= year <= 2100:
                return year
        except ValueError:
            return None
    return None


def truncate_text(value, limit=180):
    text_value = str(value or "").strip()
    if len(text_value) <= limit:
        return text_value
    return f"{text_value[:max(0, limit - 1)].rstrip()}…"


def to_max_words(value, max_words=8):
    text_value = " ".join(str(value or "").strip().split())
    if not text_value:
        return ""
    words = text_value.split(" ")
    return " ".join(words[:max_words])


def normalize_short_focus_line(value, max_words=8):
    text_value = " ".join(str(value or "").strip().split())
    if not text_value:
        return ""
    words = text_value.split(" ")
    if len(words) > max_words:
        return ""
    return text_value


def build_strategy_fallback(selected_company, combined_events, topic_counts, region_counts):
    if not combined_events:
        return {
            "investmentFocus": "Prioritize high-conviction sustainability deal themes",
            "strategicDirection": [
                f"{selected_company} has sparse recent deal signal, suggesting potential white space in themes where peers are scaling faster.",
                f"Priority is to convert selective activity into a tighter thesis linked to where {selected_company} is under-indexed versus peers.",
            ],
            "majorDeals": [],
            "yearlySummaries": {},
            "yearlyInsights": {},
        }

    sorted_topics = sorted(topic_counts.items(), key=lambda item: (-item[1], item[0].lower()))
    sorted_regions = sorted(region_counts.items(), key=lambda item: (-item[1], item[0].lower()))
    primary_topic = sorted_topics[0][0] if sorted_topics else "sustainability-aligned themes"
    primary_region = sorted_regions[0][0] if sorted_regions else "core regions"

    strategy_lines = [
        f"{selected_company} is concentrating deals in {primary_topic}, but may be leaving value on the table in adjacent sustainability themes where peers are more active.",
        f"Regional concentration in {primary_region} suggests disciplined focus, but also indicates potential missed optionality outside the core footprint.",
        f"Near-term upside is likely from sharper deal selection tied to execution gaps, not simply higher deal volume.",
    ]

    major_deals = []
    sorted_events = sorted(
        combined_events,
        key=lambda item: (item.get("year") or 0, item.get("date") or ""),
        reverse=True,
    )
    for event in sorted_events[:2]:
        major_deals.append(
            {
                "title": event.get("title") or "Key transaction",
                "date": event.get("date") or "NA",
                "source": event.get("source") or "Investment",
                "why": truncate_text(event.get("headline") or event.get("theme") or "Strategically relevant deal."),
            }
        )

    yearly = {}
    yearly_insights = {}
    events_by_year = {}
    for event in combined_events:
        year = event.get("year")
        if not year:
            continue
        events_by_year.setdefault(year, []).append(event)

    for year, events in sorted(events_by_year.items()):
        year_topics = {}
        year_regions = {}
        year_industries = {}
        for event in events:
            topic = str(event.get("theme") or "Unspecified").strip() or "Unspecified"
            region = str(event.get("region") or "Other").strip() or "Other"
            industry = str(event.get("targetIndustry") or "Other").strip() or "Other"
            year_topics[topic] = year_topics.get(topic, 0) + 1
            year_regions[region] = year_regions.get(region, 0) + 1
            year_industries[industry] = year_industries.get(industry, 0) + 1

        top_topic = sorted(year_topics.items(), key=lambda item: (-item[1], item[0].lower()))[0][0]
        top_region = sorted(year_regions.items(), key=lambda item: (-item[1], item[0].lower()))[0][0]
        year_key = str(year)
        yearly[year_key] = truncate_text(
            f"In {year}, {selected_company} concentrated on {top_topic} in {top_region}; potential value remains in under-covered themes and regions.",
            190,
        )
        top_deal_type_counts = {}
        source_counts = {"Investment Deal": 0, "Green Capex": 0}
        for event in events:
            deal_type = str(event.get("dealType") or "Other").strip() or "Other"
            top_deal_type_counts[deal_type] = top_deal_type_counts.get(deal_type, 0) + 1
            source_name = str(event.get("source") or "Investment Deal").strip()
            if source_name in source_counts:
                source_counts[source_name] += 1

        if top_deal_type_counts:
            top_deal_type, top_type_count = sorted(
                top_deal_type_counts.items(),
                key=lambda item: (-item[1], item[0].lower()),
            )[0]
        else:
            top_deal_type, top_type_count = (None, 0)
        top_type_share = (top_type_count / len(events)) if events else 0
        if top_deal_type and top_type_share >= 0.6:
            posture_line = f"{top_deal_type} dominated the playbook this year."
        elif top_deal_type:
            posture_line = f"Portfolio activity stayed mixed, led by {top_deal_type}."
        else:
            posture_line = "Portfolio activity stayed mixed across transaction types."

        if source_counts["Green Capex"] > 0 and source_counts["Investment Deal"] > 0:
            capex_line = "Execution blended investment deals with green capex scaling."
        elif source_counts["Green Capex"] > 0:
            capex_line = "Capital deployment leaned toward green capex execution."
        else:
            capex_line = "Execution leaned on transaction-led portfolio moves."

        top_industry = None
        if year_industries:
            top_industry = sorted(year_industries.items(), key=lambda item: (-item[1], item[0].lower()))[0][0]
        if top_industry and top_industry.lower() not in {"other", "-"}:
            focus_line = f"Target-industry emphasis leaned toward {top_industry}."
        else:
            focus_line = "Portfolio actions stayed spread across target industries."

        yearly_insights[year_key] = [
            truncate_text(posture_line, 120),
            truncate_text(capex_line, 120),
            truncate_text(focus_line, 120),
        ]

    return {
        "investmentFocus": "Prioritize high-conviction sustainability deal themes",
        "strategicDirection": strategy_lines,
        "majorDeals": major_deals,
        "yearlySummaries": yearly,
        "yearlyInsights": yearly_insights,
    }


def generate_investment_ai_summary(selected_company, combined_events, topic_counts, region_counts):
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    model_name = os.getenv("OPENAI_MODEL", "gpt-4o-mini").strip() or "gpt-4o-mini"

    fallback = build_strategy_fallback(selected_company, combined_events, topic_counts, region_counts)
    if not api_key or not combined_events:
        return fallback

    try:
        from openai import OpenAI

        compact_events = []
        for event in combined_events[:50]:
            compact_events.append(
                {
                    "date": event.get("date"),
                    "year": event.get("year"),
                    "title": event.get("title"),
                    "theme": event.get("theme"),
                    "driver": event.get("primaryDriver"),
                    "region": event.get("region"),
                    "source": event.get("source"),
                    "headline": event.get("headline"),
                }
            )

        client = OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model=model_name,
            temperature=0.2,
            response_format={"type": "json_object"},
            messages=[
                {
                    "role": "system",
                    "content": "You are an expert ESG investment strategy advisor. Focus on where the company is leaving value on the table vs peers. Keep outputs concise, pithy, and consulting-ready.",
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {
                            "selectedCompany": selected_company,
                            "events": compact_events,
                            "required": {
                                "investmentFocus": "single line, max 8 words, punchy and specific to current investment direction, no trailing conjunctions",
                                "strategicDirection": "array of exactly 2-3 concise strings about missed value opportunities and strategic focus",
                                "majorDeals": "array of 1-2 objects with title,date,source,why where why explains strategy significance without using deal value",
                                "yearlySummaries": "object of year->1 concise sentence on strategy and value-left-on-table",
                                "yearlyInsights": "object of year->array of exactly 3 short, pithy bullet points (max 12 words each); avoid repeating raw deal counts and avoid theme/region callouts",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ],
        )

        content = response.choices[0].message.content or "{}"
        parsed = json.loads(content)

        investment_focus = normalize_short_focus_line(parsed.get("investmentFocus") or "", 8)
        if not investment_focus:
            investment_focus = fallback["investmentFocus"]

        strategic_direction = [
            truncate_text(item, 200)
            for item in (parsed.get("strategicDirection") or [])
            if str(item or "").strip()
        ]

        major_deals = []
        for item in (parsed.get("majorDeals") or []):
            if not isinstance(item, dict):
                continue
            title = str(item.get("title") or "").strip()
            why = str(item.get("why") or "").strip()
            if not title or not why:
                continue
            major_deals.append(
                {
                    "title": truncate_text(title, 90),
                    "date": str(item.get("date") or "NA").strip(),
                    "source": str(item.get("source") or "Investment").strip(),
                    "why": truncate_text(why, 220),
                }
            )

        yearly_summaries_raw = parsed.get("yearlySummaries") or {}
        yearly_summaries = {}
        if isinstance(yearly_summaries_raw, dict):
            for year_key, summary in yearly_summaries_raw.items():
                year_text = str(year_key or "").strip()
                summary_text = str(summary or "").strip()
                if year_text and summary_text:
                    yearly_summaries[year_text] = truncate_text(summary_text, 220)

        yearly_insights_raw = parsed.get("yearlyInsights") or {}
        yearly_insights = {}
        if isinstance(yearly_insights_raw, dict):
            for year_key, points in yearly_insights_raw.items():
                year_text = str(year_key or "").strip()
                if not year_text or not isinstance(points, list):
                    continue
                cleaned_points = [
                    truncate_text(str(point or "").strip(), 120)
                    for point in points
                    if str(point or "").strip()
                ][:3]
                if cleaned_points:
                    yearly_insights[year_text] = cleaned_points

        if len(strategic_direction) < 2:
            strategic_direction = fallback["strategicDirection"]
        if not major_deals:
            major_deals = fallback["majorDeals"]
        if not yearly_summaries:
            yearly_summaries = fallback["yearlySummaries"]
        if not yearly_insights:
            yearly_insights = fallback.get("yearlyInsights", {})

        return {
            "investmentFocus": investment_focus,
            "strategicDirection": strategic_direction[:3],
            "majorDeals": major_deals[:2],
            "yearlySummaries": yearly_summaries,
            "yearlyInsights": yearly_insights,
        }
    except Exception:
        return fallback


def generate_pithy_focus_company_insights(selected_company, insights):
    raw_insights = [str(item or "").strip() for item in (insights or []) if str(item or "").strip()]
    if not raw_insights:
        return {
            "insights": [],
            "source": "fallback",
            "reason": "no_input_insights",
            "model": None,
            "hasApiKey": bool(os.getenv("OPENAI_API_KEY", "").strip()),
        }

    fallback_points = [truncate_text(point, 120) for point in raw_insights[:3]]

    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    model_name = os.getenv("OPENAI_MODEL", "gpt-4o-mini").strip() or "gpt-4o-mini"
    if not api_key:
        return {
            "insights": fallback_points,
            "source": "fallback",
            "reason": "missing_openai_api_key",
            "model": model_name,
            "hasApiKey": False,
        }

    try:
        from openai import OpenAI

        client = OpenAI(api_key=api_key)
        prompt_payload = {
            "selectedCompany": selected_company,
            "focusCompanyInsights": raw_insights[:3],
            "required": {
                "pithyBullets": "array of exactly 3 bullets, each <= 12 words, specific and actionable"
            },
        }

        def _extract_bullets(text):
            lines = [str(line or "").strip() for line in str(text or "").splitlines()]
            cleaned = []
            for line in lines:
                if not line:
                    continue
                normalized = line
                if normalized[:1] in {"-", "•", "*"}:
                    normalized = normalized[1:].strip()
                if ". " in normalized[:4] and normalized[:1].isdigit():
                    normalized = normalized.split(". ", 1)[1].strip()
                if normalized:
                    cleaned.append(truncate_text(normalized, 120))
            return cleaned[:3]

        try:
            response = client.chat.completions.create(
                model=model_name,
                temperature=0.2,
                response_format={"type": "json_object"},
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert ESG advisor. Rewrite insights into pithy executive bullets with no fluff.",
                    },
                    {
                        "role": "user",
                        "content": json.dumps(prompt_payload, ensure_ascii=False),
                    },
                ],
            )

            content = response.choices[0].message.content or "{}"
            parsed = json.loads(content)
            bullets = [
                truncate_text(str(point or "").strip(), 120)
                for point in (parsed.get("pithyBullets") or [])
                if str(point or "").strip()
            ]

            if len(bullets) >= 3:
                return {
                    "insights": bullets[:3],
                    "source": "ai",
                    "reason": "openai_success",
                    "model": model_name,
                    "hasApiKey": True,
                }
        except Exception:
            pass

        try:
            response = client.chat.completions.create(
                model=model_name,
                temperature=0.2,
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert ESG advisor. Return exactly 3 short bullets, one per line.",
                    },
                    {
                        "role": "user",
                        "content": (
                            f"Company: {selected_company}\n"
                            f"Insights:\n- {raw_insights[0]}\n- {raw_insights[1] if len(raw_insights) > 1 else ''}\n- {raw_insights[2] if len(raw_insights) > 2 else ''}\n"
                            "Rewrite as 3 pithy executive bullets, max 12 words each."
                        ),
                    },
                ],
            )

            plain_content = response.choices[0].message.content or ""
            plain_bullets = _extract_bullets(plain_content)
            if len(plain_bullets) >= 3:
                return {
                    "insights": plain_bullets[:3],
                    "source": "ai",
                    "reason": "openai_success_plaintext",
                    "model": model_name,
                    "hasApiKey": True,
                }
        except Exception as plain_error:
            return {
                "insights": fallback_points,
                "source": "fallback",
                "reason": f"openai_error:{type(plain_error).__name__}",
                "model": model_name,
                "hasApiKey": True,
            }

        return {
            "insights": fallback_points,
            "source": "fallback",
            "reason": "invalid_ai_payload",
            "model": model_name,
            "hasApiKey": True,
        }
    except Exception as error:
        return {
            "insights": fallback_points,
            "source": "fallback",
            "reason": f"openai_error:{type(error).__name__}",
            "model": model_name,
            "hasApiKey": True,
        }


def generate_pithy_peer_insights(selected_company, peer_insights):
    peer_context_name = f"{selected_company} peer set"
    return generate_pithy_focus_company_insights(peer_context_name, peer_insights)


def build_peer_focus_bullets(selected_company, selected_topic_totals, peer_topic_totals, peer_count):
    safe_peer_count = max(1, int(peer_count or 1))

    selected_sorted = sorted(
        selected_topic_totals.items(),
        key=lambda item: (-item[1], item[0].lower()),
    )
    peer_avg_map = {
        topic: (float(count) / safe_peer_count)
        for topic, count in peer_topic_totals.items()
    }
    peer_sorted = sorted(peer_avg_map.items(), key=lambda item: (-item[1], item[0].lower()))

    selected_top = selected_sorted[0][0] if selected_sorted else "sustainable operations"
    peer_top = peer_sorted[0][0] if peer_sorted else "diversified sustainability themes"

    bullets = []
    if selected_top.lower() == peer_top.lower():
        bullets.append(
            truncate_text(
                f"{selected_company}: stay focused on {selected_top}, but tighten the deal filter to fewer, higher-impact bets with clear delivery owners.",
                220,
            )
        )
    else:
        bullets.append(
            truncate_text(
                f"{selected_company}: keep {selected_top} as core, and launch 1-2 pilot deals in {peer_top} to close strategic white space.",
                220,
            )
        )

    union_topics = sorted(set(selected_topic_totals.keys()) | set(peer_avg_map.keys()))
    positive_gaps = []
    for topic in union_topics:
        selected_value = float(selected_topic_totals.get(topic, 0))
        peer_value = float(peer_avg_map.get(topic, 0))
        gap = peer_value - selected_value
        positive_gaps.append((topic, gap))

    positive_gaps.sort(key=lambda item: item[1], reverse=True)
    top_gap_topic, top_gap_value = positive_gaps[0] if positive_gaps else ("", 0)
    if top_gap_topic and top_gap_value > 0:
        bullets.append(
            truncate_text(
                f"Build a 12-month target list in {top_gap_topic} and pursue one partnership or bolt-on to close the peer gap quickly.",
                220,
            )
        )
    else:
        bullets.append(
            truncate_text(
                f"Convert focus into outcomes: require every deal to pass value gates on growth, risk reduction, and sustainability impact.",
                220,
            )
        )

    bullets.append(
        truncate_text(
            f"Set quarterly execution reviews so {selected_company} can rebalance capital faster when themes underperform.",
            220,
        )
    )

    return bullets[:3]


def build_investment_insights(sector_name, selected_company, sector_companies):
    investment_deals = load_investment_deals_dataset()
    green_capex_deals = load_green_capex_dataset()
    rationals_dataset = load_investment_rationals_dataset()

    def matches_selected_company(row):
        master_company = row.get("Master Company") or ""
        buyers = row.get("Buyers/Investors") or ""
        return company_text_matches(master_company, selected_company) or company_text_matches(buyers, selected_company)

    selected_deals = [row for row in investment_deals if isinstance(row, dict) and matches_selected_company(row)]

    selected_capex = [
        row
        for row in green_capex_deals
        if isinstance(row, dict)
        and (
            company_text_matches(row.get("Buyers/Investors") or "", selected_company)
            or company_text_matches(row.get("Master Company") or "", selected_company)
        )
    ]

    def is_sustainability_classification(theme_value):
        theme_text = str(theme_value or "").strip()
        return bool(theme_text) and theme_text.lower() != "none"

    closed_deal_count = sum(
        1
        for row in selected_deals
        if str(row.get("Transaction Status") or "").strip().lower() == "closed"
    )

    def top_counts(rows, field_name, limit=5):
        counter = {}
        for row in rows:
            value = str(row.get(field_name) or "").strip()
            if not value:
                continue
            counter[value] = counter.get(value, 0) + 1
        return [
            {"label": item[0], "count": item[1]}
            for item in sorted(counter.items(), key=lambda pair: (-pair[1], pair[0].lower()))[:limit]
        ]

    def get_row_theme(row):
        return str(row.get("classification_L1") or row.get("classification") or "").strip()

    def get_row_target_industry(row):
        return str(row.get("Master Primary Industry Target") or row.get("Primary Industry [Target/Issuer]") or "").strip()

    def aggregate_theme_counts(deals, capex):
        counter = {}
        for row in deals:
            theme = get_row_theme(row)
            if not is_sustainability_classification(theme):
                continue
            counter[theme] = counter.get(theme, 0) + 1

        for row in capex:
            theme = str(row.get("classification") or row.get("classification_L1") or "").strip()
            if not is_sustainability_classification(theme):
                continue
            counter[theme] = counter.get(theme, 0) + 1
        return counter

    def aggregate_region_counts(deals):
        counter = {}
        for row in deals:
            if not is_sustainability_classification(get_row_theme(row)):
                continue
            region = str(row.get("Master Target Region") or "").strip() or "Other"
            counter[region] = counter.get(region, 0) + 1
        return counter

    selected_sustainability_deals = [row for row in selected_deals if is_sustainability_classification(get_row_theme(row))]
    selected_sustainability_capex = [
        row
        for row in selected_capex
        if is_sustainability_classification(str(row.get("classification") or row.get("classification_L1") or "").strip())
    ]

    recent_deals = []
    for row in selected_sustainability_deals:
        announced_date_raw = row.get("All Transactions Announced Date") or row.get("Year")
        announced_date = parse_flexible_date(announced_date_raw)
        announced_year = parse_year(announced_date_raw)

        overview_items = row.get("Overview") or row.get("Why_The_Deal_Happened") or []
        overview_points = [str(item).strip() for item in overview_items if str(item or "").strip()] if isinstance(overview_items, list) else []

        headline = overview_points[0] if overview_points else ""
        if not headline:
            headline = str(row.get("Primary Driver Justification") or "").strip()
        if not headline:
            headline = str(row.get("Transaction Comments") or "").strip().split("\n")[0]

        raw_date_text = str(announced_date_raw or "").strip()
        looks_like_year_only = raw_date_text.isdigit() and len(raw_date_text) == 4
        if looks_like_year_only:
            date_label = str(announced_year or raw_date_text)
        else:
            date_label = announced_date.strftime("%Y-%m-%d") if announced_date else str(announced_year or announced_date_raw or "NA")

        recent_deals.append(
            {
                "date": date_label,
                "_sortDate": announced_date or datetime.min,
                "year": announced_year,
                "target": str(row.get("Master Target") or row.get("Target/Issuer") or "Unknown target").strip(),
                "transactionType": str(row.get("Transaction Types") or "").strip(),
                "transactionValue": str(row.get("Total Transaction Value ($USDmm, Historical rate)") or "").strip(),
                "targetIndustry": get_row_target_industry(row),
                "classification": get_row_theme(row),
                "primaryDriver": str(row.get("Primary Driver") or "").strip(),
                "region": str(row.get("Master Target Region") or "").strip() or "Other",
                "headline": headline,
                "overviewPoints": overview_points,
            }
        )

    recent_deals.sort(key=lambda item: item["_sortDate"], reverse=True)
    for item in recent_deals:
        item.pop("_sortDate", None)

    peer_benchmark = []
    peer_topic_totals = {}
    peer_region_totals = {}
    selected_topic_totals = aggregate_theme_counts(selected_deals, selected_capex)
    selected_region_totals = aggregate_region_counts(selected_deals)

    for company_name in sector_companies:
        company_deals = [
            row
            for row in investment_deals
            if isinstance(row, dict)
            and (
                company_text_matches(row.get("Master Company") or "", company_name)
                or company_text_matches(row.get("Buyers/Investors") or "", company_name)
            )
        ]
        company_capex = [
            row
            for row in green_capex_deals
            if isinstance(row, dict)
            and (
                company_text_matches(row.get("Buyers/Investors") or "", company_name)
                or company_text_matches(row.get("Master Company") or "", company_name)
            )
        ]
        company_theme_totals = aggregate_theme_counts(company_deals, company_capex)
        company_region_totals = aggregate_region_counts(company_deals)

        if normalize_company_name(company_name) != normalize_company_name(selected_company):
            for key, value in company_theme_totals.items():
                peer_topic_totals[key] = peer_topic_totals.get(key, 0) + value
            for key, value in company_region_totals.items():
                peer_region_totals[key] = peer_region_totals.get(key, 0) + value

        peer_benchmark.append(
            {
                "company": company_name,
                "isSelected": normalize_company_name(company_name) == normalize_company_name(selected_company),
                "dealCount": len(company_deals),
                "greenCapexCount": len(company_capex),
            }
        )

    peer_benchmark.sort(key=lambda item: (-item["dealCount"], item["company"].lower()))
    selected_benchmark = next((item for item in peer_benchmark if item.get("isSelected")), None)
    sorted_peers = [item for item in peer_benchmark if not item.get("isSelected")]
    if selected_benchmark:
        peer_benchmark = [selected_benchmark, *sorted_peers]

    capex_highlights = []
    for row in selected_sustainability_capex:
        capex_date_raw = row.get("All Transactions Announced Date")
        capex_date = parse_flexible_date(capex_date_raw)
        capex_year = parse_year(capex_date_raw)
        capex_highlights.append(
            {
                "date": capex_date.strftime("%Y-%m-%d") if capex_date else str(capex_date_raw or "NA"),
                "_sortDate": capex_date or datetime.min,
                "year": capex_year,
                "initiative": str(row.get("Target/Issuer") or "").strip(),
                "category": str(row.get("Green Investment Category") or "").strip(),
                "dealType": "Green Capex",
                "targetIndustry": str(row.get("Primary Industry [Target/Issuer]") or row.get("Master Primary Industry Target") or "").strip(),
                "classification": str(row.get("classification") or "").strip(),
                "region": str(row.get("Master Target Region") or "").strip() or "Other",
                "excerpt": str(row.get("Source Excerpt") or row.get("AI Interpretation") or "").strip(),
                "value": str(row.get("Transaction Value") or "").strip(),
                "unit": str(row.get("Transaction Unit") or "").strip(),
                "overviewPoints": [str(row.get("Source Excerpt") or row.get("AI Interpretation") or "").strip()],
            }
        )

    capex_highlights.sort(key=lambda item: item["_sortDate"], reverse=True)
    for item in capex_highlights:
        item.pop("_sortDate", None)

    selected_rationale = next(
        (
            row for row in rationals_dataset
            if isinstance(row, dict) and company_text_matches(row.get("Focus Company") or "", selected_company)
        ),
        None,
    )
    focus_company_insights = (selected_rationale or {}).get("FC Insights", []) if isinstance(selected_rationale, dict) else []
    peer_insights = (selected_rationale or {}).get("Peers Insights", []) if isinstance(selected_rationale, dict) else []
    pithy_focus_company_result = generate_pithy_focus_company_insights(selected_company, focus_company_insights)
    pithy_focus_company_insights = pithy_focus_company_result.get("insights", [])
    pithy_peer_result = generate_pithy_peer_insights(selected_company, peer_insights)
    pithy_peer_insights = pithy_peer_result.get("insights", [])

    all_regions_selected = aggregate_region_counts(selected_deals)
    total_region_events_selected = sum(all_regions_selected.values())
    top_region = None
    top_region_share_pct = 0.0
    if total_region_events_selected > 0:
        top_region = sorted(all_regions_selected.items(), key=lambda item: (-item[1], item[0].lower()))[0][0]
        top_region_count = all_regions_selected[top_region]
        top_region_share_pct = round((top_region_count / total_region_events_selected) * 100, 1)

    filtered_2526_topics = {}
    for deal in recent_deals:
        year = deal.get("year")
        theme = str(deal.get("classification") or "").strip()
        if year in {2025, 2026} and is_sustainability_classification(theme):
            filtered_2526_topics[theme] = filtered_2526_topics.get(theme, 0) + 1

    for capex in capex_highlights:
        year = capex.get("year")
        theme = str(capex.get("classification") or "").strip()
        if year in {2025, 2026} and is_sustainability_classification(theme):
            filtered_2526_topics[theme] = filtered_2526_topics.get(theme, 0) + 1

    def build_sustainability_trend(selected_name, peers):
        selected_trend_counts = {}
        peer_trend_counts_by_company = {}

        def update_counts(bucket, year, total_increment, sustainability_increment):
            year_key = int(year)
            if year_key not in bucket:
                bucket[year_key] = {"total": 0, "sustainability": 0}
            bucket[year_key]["total"] += total_increment
            bucket[year_key]["sustainability"] += sustainability_increment

        peer_names = [
            peer_name
            for peer_name in peers
            if normalize_company_name(peer_name) != normalize_company_name(selected_name)
        ]

        for row in investment_deals:
            if not isinstance(row, dict):
                continue

            year = parse_year(row.get("All Transactions Announced Date") or row.get("Year"))
            if not year:
                continue

            row_company = str(row.get("Master Company") or "")
            row_buyers = str(row.get("Buyers/Investors") or "")
            is_selected = company_text_matches(row_company, selected_name) or company_text_matches(row_buyers, selected_name)
            sustainability_flag = 1 if is_sustainability_classification(get_row_theme(row)) else 0

            if is_selected:
                update_counts(selected_trend_counts, year, 1, sustainability_flag)

            for peer_name in peer_names:
                if company_text_matches(row_company, peer_name) or company_text_matches(row_buyers, peer_name):
                    peer_bucket = peer_trend_counts_by_company.setdefault(peer_name, {})
                    update_counts(peer_bucket, year, 1, sustainability_flag)

        all_peer_years = set()
        for peer_year_map in peer_trend_counts_by_company.values():
            all_peer_years.update(peer_year_map.keys())

        all_years = sorted(set(selected_trend_counts.keys()) | all_peer_years)
        if all_years:
            all_years = all_years[-3:]

        trend_rows = []
        for year in all_years:
            selected_totals = selected_trend_counts.get(year, {"total": 0, "sustainability": 0})
            selected_pct = 0.0
            if selected_totals["total"] > 0:
                selected_pct = (selected_totals["sustainability"] / selected_totals["total"]) * 100

            peer_pcts = []
            for peer_year_map in peer_trend_counts_by_company.values():
                peer_year_totals = peer_year_map.get(year, {"total": 0, "sustainability": 0})
                if peer_year_totals["total"] > 0:
                    peer_pcts.append((peer_year_totals["sustainability"] / peer_year_totals["total"]) * 100)

            peer_pct = (sum(peer_pcts) / len(peer_pcts)) if peer_pcts else 0.0

            trend_rows.append(
                {
                    "year": str(year),
                    "sustainabilityPct": round(selected_pct, 1),
                    "peerAvgPct": round(peer_pct, 1),
                }
            )

        return trend_rows

    sustainability_trend = build_sustainability_trend(selected_company, sector_companies)

    top_topics_2526 = [
        {"label": key, "count": value}
        for key, value in sorted(filtered_2526_topics.items(), key=lambda item: (-item[1], item[0].lower()))[:3]
    ]

    all_topic_keys = sorted([key for key in selected_topic_totals.keys() if str(key or "").strip()])
    all_region_keys = sorted([key for key in selected_region_totals.keys() if str(key or "").strip()])

    selected_peer_count = max(1, len([name for name in sector_companies if normalize_company_name(name) != normalize_company_name(selected_company)]))

    topic_spider = [
        {
            "theme": topic_key,
            "companyScore": float(selected_topic_totals.get(topic_key, 0)),
            "peerAvg": round(float(peer_topic_totals.get(topic_key, 0)) / selected_peer_count, 2),
        }
        for topic_key in all_topic_keys
    ]
    topic_spider.sort(key=lambda item: item["companyScore"], reverse=True)

    region_spider = [
        {
            "theme": region_key,
            "companyScore": float(selected_region_totals.get(region_key, 0)),
            "peerAvg": round(float(peer_region_totals.get(region_key, 0)) / selected_peer_count, 2),
        }
        for region_key in all_region_keys
    ]
    region_spider.sort(key=lambda item: item["companyScore"], reverse=True)

    top_topic_keys = [item["theme"] for item in topic_spider[:3]]
    if not top_topic_keys:
        top_topic_keys = ["Unspecified"]

    top_region_keys = [item["theme"] for item in region_spider[:4]]
    if not top_region_keys:
        top_region_keys = ["Other"]

    topic_breakdown_rows = []
    region_breakdown_rows = []
    target_industry_breakdown_rows = []
    deal_type_breakdown_rows = []

    def normalize_deal_type(value):
        text = str(value or "").strip().lower()
        if not text or text == "-":
            return "Other"
        if "minority investment" in text:
            return "Minority Investment"
        if "majority investment" in text or "control acquisition" in text:
            return "Majority Investment"
        if "full acquisition" in text:
            return "Full Acquisition"
        if "divestment" in text or "exit" in text:
            return "Divestment / Exit"
        if "spin-off" in text or "spin off" in text or "split-off" in text or "demerger" in text:
            return "Spin-off / Demerger"
        if "sponsor" in text:
            return "Sponsor"
        if "strategic" in text:
            return "Strategic"
        return "Other"

    def normalize_target_industry(value):
        text = str(value or "").strip()
        if not text or text == "-":
            return "Other"
        if text.lower() == "consumer":
            return "Consumer Products"
        return text

    company_target_industry_counts = {}
    company_deal_type_counts = {}

    for company_name in sector_companies:
        company_deals = [
            row
            for row in investment_deals
            if isinstance(row, dict)
            and (
                company_text_matches(row.get("Master Company") or "", company_name)
                or company_text_matches(row.get("Buyers/Investors") or "", company_name)
            )
        ]
        company_capex = [
            row
            for row in green_capex_deals
            if isinstance(row, dict)
            and (
                company_text_matches(row.get("Buyers/Investors") or "", company_name)
                or company_text_matches(row.get("Master Company") or "", company_name)
            )
        ]
        company_theme_counts = aggregate_theme_counts(company_deals, company_capex)
        company_region_counts = aggregate_region_counts(company_deals)

        company_target_counts = {}
        company_deal_counts = {}
        for row in company_deals:
            if not isinstance(row, dict):
                continue
            target_industry = normalize_target_industry(row.get("Master Primary Industry Target"))
            company_target_counts[target_industry] = company_target_counts.get(target_industry, 0) + 1

            deal_type = normalize_deal_type(row.get("Transaction Types") or row.get("Transaction Type"))
            company_deal_counts[deal_type] = company_deal_counts.get(deal_type, 0) + 1

        company_target_industry_counts[company_name] = company_target_counts
        company_deal_type_counts[company_name] = company_deal_counts

        topic_row = {
            "company": company_name,
            "isSelected": normalize_company_name(company_name) == normalize_company_name(selected_company),
        }
        topic_other = 0
        for key, value in company_theme_counts.items():
            if key in top_topic_keys:
                topic_row[key] = value
            else:
                topic_other += value
        topic_row["Other"] = topic_other
        topic_breakdown_rows.append(topic_row)

        region_row = {
            "company": company_name,
            "isSelected": normalize_company_name(company_name) == normalize_company_name(selected_company),
        }
        region_other = 0
        for key, value in company_region_counts.items():
            if key in top_region_keys:
                region_row[key] = value
            else:
                region_other += value
        region_row["Other"] = region_other
        region_breakdown_rows.append(region_row)

    aggregated_target_industry_counts = {}
    aggregated_deal_type_counts = {}
    for company_name in sector_companies:
        for key, value in (company_target_industry_counts.get(company_name) or {}).items():
            aggregated_target_industry_counts[key] = aggregated_target_industry_counts.get(key, 0) + value
        for key, value in (company_deal_type_counts.get(company_name) or {}).items():
            aggregated_deal_type_counts[key] = aggregated_deal_type_counts.get(key, 0) + value

    top_target_industry_keys = [
        key
        for key, _ in sorted(
            aggregated_target_industry_counts.items(),
            key=lambda item: (-item[1], item[0].lower()),
        )[:4]
    ]
    if not top_target_industry_keys:
        top_target_industry_keys = ["Other"]

    top_deal_type_keys = [
        key
        for key, _ in sorted(
            aggregated_deal_type_counts.items(),
            key=lambda item: (-item[1], item[0].lower()),
        )[:4]
    ]
    if not top_deal_type_keys:
        top_deal_type_keys = ["Other"]

    for company_name in sector_companies:
        target_counts = company_target_industry_counts.get(company_name, {})
        deal_counts = company_deal_type_counts.get(company_name, {})

        target_row = {
            "company": company_name,
            "isSelected": normalize_company_name(company_name) == normalize_company_name(selected_company),
        }
        target_other = 0
        for key, value in target_counts.items():
            if key in top_target_industry_keys:
                target_row[key] = value
            else:
                target_other += value
        target_row["Other"] = target_other
        target_industry_breakdown_rows.append(target_row)

        deal_row = {
            "company": company_name,
            "isSelected": normalize_company_name(company_name) == normalize_company_name(selected_company),
        }
        deal_other = 0
        for key, value in deal_counts.items():
            if key in top_deal_type_keys:
                deal_row[key] = value
            else:
                deal_other += value
        deal_row["Other"] = deal_other
        deal_type_breakdown_rows.append(deal_row)

    topic_breakdown_rows.sort(key=lambda item: (not item["isSelected"], item["company"].lower()))
    region_breakdown_rows.sort(key=lambda item: (not item["isSelected"], item["company"].lower()))
    target_industry_breakdown_rows.sort(key=lambda item: (not item["isSelected"], item["company"].lower()))
    deal_type_breakdown_rows.sort(key=lambda item: (not item["isSelected"], item["company"].lower()))

    timeline_events = []
    for deal in recent_deals:
        timeline_events.append(
            {
                "source": "Investment Deal",
                "date": deal.get("date"),
                "year": deal.get("year"),
                "title": deal.get("target"),
                "theme": deal.get("classification") or "Unspecified",
                "targetIndustry": deal.get("targetIndustry") or "Other",
                "dealType": deal.get("transactionType") or "Investment Deal",
                "transactionValue": deal.get("transactionValue") or "",
                "region": deal.get("region") or "Other",
                "primaryDriver": deal.get("primaryDriver"),
                "headline": deal.get("headline"),
                "overviewPoints": deal.get("overviewPoints") or [],
            }
        )

    for capex in capex_highlights:
        timeline_events.append(
            {
                "source": "Green Capex",
                "date": capex.get("date"),
                "year": capex.get("year"),
                "title": capex.get("initiative") or "Capex initiative",
                "theme": capex.get("classification") or capex.get("category") or "Unspecified",
                "targetIndustry": capex.get("targetIndustry") or capex.get("category") or "Operations / Infrastructure",
                "dealType": capex.get("dealType") or "Green Capex",
                "transactionValue": " ".join(
                    part
                    for part in [str(capex.get("value") or "").strip(), str(capex.get("unit") or "").strip()]
                    if part
                ),
                "region": capex.get("region") or "Other",
                "primaryDriver": "Sustainability / operations",
                "headline": capex.get("excerpt"),
                "overviewPoints": [point for point in (capex.get("overviewPoints") or []) if point],
            }
        )

    timeline_events = [event for event in timeline_events if event.get("year")]
    timeline_events.sort(key=lambda item: (item.get("year") or 0, item.get("date") or ""))

    all_timeline_events_for_strategy = list(timeline_events)

    years_sorted = sorted({event["year"] for event in timeline_events})
    if years_sorted:
        latest_year = years_sorted[-1]
        recent_years = {year for year in years_sorted if year >= latest_year - 2}
        timeline_events = [event for event in timeline_events if event["year"] in recent_years]

    ai_strategy = generate_investment_ai_summary(selected_company, all_timeline_events_for_strategy, selected_topic_totals, selected_region_totals)
    difference_bullets = build_peer_focus_bullets(
        selected_company,
        selected_topic_totals,
        peer_topic_totals,
        selected_peer_count,
    )

    region_shares = []
    if total_region_events_selected > 0:
        sorted_regions = sorted(all_regions_selected.items(), key=lambda item: (-item[1], item[0].lower()))
        for region_name, region_count in sorted_regions:
            region_shares.append(
                {
                    "region": region_name,
                    "sharePct": round((region_count / total_region_events_selected) * 100, 1),
                }
            )

    top_topics_labels = [item["label"] for item in top_topics_2526[:2]]

    timeline_by_year = {}
    for event in timeline_events:
        year_key = str(event.get("year"))
        timeline_by_year.setdefault(year_key, []).append(event)

    def _top_with_share(counter_map):
        if not counter_map:
            return None, 0, 0.0
        total = sum(counter_map.values())
        top_key, top_count = sorted(counter_map.items(), key=lambda item: (-item[1], item[0].lower()))[0]
        share = (top_count / total) if total else 0.0
        return top_key, top_count, share

    def _deal_type_strategy_phrase(deal_type):
        label = str(deal_type or "").strip().lower()
        if label in {"full acquisition", "majority investment", "sponsor", "strategic"}:
            return "the year prioritized control-oriented portfolio expansion"
        if label in {"divestment / exit", "spin-off / demerger"}:
            return "the year emphasized portfolio pruning and capital recycling"
        if label == "minority investment":
            return "the year leaned toward option-value bets and capability scouting"
        return "the year maintained a mixed portfolio posture"

    def build_year_overview_points(events, previous_events=None):
        event_count = len(events)
        if event_count == 0:
            return ["No yearly strategy summary available."]

        deal_type_counts = {}
        theme_counts = {}
        region_counts = {}
        industry_counts = {}

        for event in events:
            deal_type = str(event.get("dealType") or "Other").strip() or "Other"
            theme = str(event.get("theme") or "Unspecified").strip() or "Unspecified"
            region = str(event.get("region") or "Other").strip() or "Other"
            industry = str(event.get("targetIndustry") or "Other").strip() or "Other"

            deal_type_counts[deal_type] = deal_type_counts.get(deal_type, 0) + 1
            theme_counts[theme] = theme_counts.get(theme, 0) + 1
            region_counts[region] = region_counts.get(region, 0) + 1
            industry_counts[industry] = industry_counts.get(industry, 0) + 1

        top_deal_type, _, top_deal_type_share = _top_with_share(deal_type_counts)
        top_industry, _, _ = _top_with_share(industry_counts)

        if top_deal_type_share >= 0.6:
            posture_phrase = f"{top_deal_type} dominated the yearly playbook"
        else:
            posture_phrase = f"activity stayed mixed, led by {top_deal_type}"

        strategy_phrase = _deal_type_strategy_phrase(top_deal_type)

        source_counts = {"Investment Deal": 0, "Green Capex": 0}
        for event in events:
            source_name = str(event.get("source") or "Investment Deal").strip()
            if source_name in source_counts:
                source_counts[source_name] += 1

        if source_counts["Investment Deal"] > 0 and source_counts["Green Capex"] > 0:
            execution_phrase = "investment deals and green capex progressed in parallel"
        elif source_counts["Green Capex"] > 0:
            execution_phrase = "capital deployment leaned toward green capex implementation"
        else:
            execution_phrase = "capital deployment was driven by transaction execution"

        yoy_phrase = ""
        if previous_events is not None:
            previous_count = len(previous_events)
            if previous_count > 0:
                delta = event_count - previous_count
                if delta > 0:
                    yoy_phrase = "deal velocity accelerated versus the prior year"
                elif delta < 0:
                    yoy_phrase = "deal velocity softened versus the prior year"
                else:
                    yoy_phrase = "deal velocity held steady versus the prior year"

        industry_phrase = ""
        if top_industry and top_industry.lower() not in {"other", "-"}:
            industry_phrase = f"target focus leaned toward {top_industry}"

        points = [
            truncate_text(posture_phrase + ".", 130),
            truncate_text(execution_phrase + ".", 130),
            truncate_text((industry_phrase or strategy_phrase) + (f"; {yoy_phrase}." if yoy_phrase else "."), 150),
        ]

        return points[:3]

    timeline_years = []
    previous_year_events = None
    for year_key in sorted(timeline_by_year.keys()):
        events = sorted(timeline_by_year[year_key], key=lambda item: item.get("date") or "")
        ai_points = ai_strategy.get("yearlyInsights", {}).get(str(year_key), []) if isinstance(ai_strategy, dict) else []
        fallback_points = build_year_overview_points(events, previous_year_events)
        year_points = [
            truncate_text(str(point or "").strip(), 140)
            for point in (ai_points if isinstance(ai_points, list) else [])
            if str(point or "").strip()
        ][:3]
        filtered_year_points = []
        for point in year_points:
            point_lc = point.lower()
            low_value_pattern = (
                " total deal" in point_lc
                or " deals executed" in point_lc
                or "theme concentration" in point_lc
                or "regional focus" in point_lc
                or " lead region" in point_lc
            )
            if not low_value_pattern:
                filtered_year_points.append(point)
        year_points = filtered_year_points
        if len(year_points) < 3:
            for point in fallback_points:
                if len(year_points) >= 3:
                    break
                if point not in year_points:
                    year_points.append(point)

        timeline_years.append(
            {
                "year": year_key,
                "dealCount": len(events),
                "summary": " ".join(year_points),
                "insightPoints": year_points[:3],
                "events": events,
            }
        )
        previous_year_events = events

    latest_date = None
    if recent_deals:
        latest_date = recent_deals[0].get("date")

    return {
        "sector": sector_name,
        "selectedCompany": selected_company,
        "summary": {
            "dealCount": len(selected_sustainability_deals),
            "closedDealCount": closed_deal_count,
            "greenCapexCount": len(selected_sustainability_capex),
            "totalCompanyDeals": len(selected_deals),
            "latestDealDate": latest_date,
            "topRegion": top_region,
            "topRegionSharePct": top_region_share_pct,
        },
        "topClassifications": top_counts(
            [{**row, "classification": get_row_theme(row)} for row in selected_sustainability_deals],
            "classification",
            limit=4,
        ),
        "topDrivers": top_counts(selected_sustainability_deals, "Primary Driver", limit=4),
        "topBuckets": top_counts(selected_sustainability_deals, "Primary Bucket", limit=4),
        "investmentFocus": ai_strategy.get("investmentFocus") or "Refocus portfolio toward resilient sustainability bets",
        "regionShares": region_shares,
        "topSustainabilityTopics": top_topics_labels,
        "topTopics2025_26": top_topics_2526,
        "greenCapexCategories": top_counts(selected_sustainability_capex, "Green Investment Category", limit=5),
        "sustainabilityTrend": sustainability_trend,
        "peerBenchmark": peer_benchmark,
        "recentDeals": recent_deals[:10],
        "greenCapexHighlights": capex_highlights[:8],
        "differenceBullets": difference_bullets[:3],
        "strategicDirection": {
            "summaryLines": ai_strategy.get("strategicDirection", []),
            "majorDeals": ai_strategy.get("majorDeals", []),
        },
        "spider": {
            "topics": topic_spider[:8],
            "regions": region_spider[:8],
        },
        "charts": {
            "topics": {
                "keys": [*top_topic_keys, "Other"],
                "rows": topic_breakdown_rows,
            },
            "regions": {
                "keys": [*top_region_keys, "Other"],
                "rows": region_breakdown_rows,
            },
            "targetIndustries": {
                "keys": [*top_target_industry_keys, "Other"],
                "rows": target_industry_breakdown_rows,
            },
            "dealTypes": {
                "keys": [*top_deal_type_keys, "Other"],
                "rows": deal_type_breakdown_rows,
            },
        },
        "timeline": {
            "years": timeline_years,
        },
        "narrative": {
            "focusCompanyInsights": focus_company_insights,
            "focusCompanyPithyInsights": pithy_focus_company_insights,
            "focusCompanyPithyMeta": {
                "source": pithy_focus_company_result.get("source", "fallback"),
                "reason": pithy_focus_company_result.get("reason", "unknown"),
                "model": pithy_focus_company_result.get("model"),
                "hasApiKey": bool(pithy_focus_company_result.get("hasApiKey", False)),
            },
            "peerInsights": peer_insights,
            "peerPithyInsights": pithy_peer_insights,
            "peerPithyMeta": {
                "source": pithy_peer_result.get("source", "fallback"),
                "reason": pithy_peer_result.get("reason", "unknown"),
                "model": pithy_peer_result.get("model"),
                "hasApiKey": bool(pithy_peer_result.get("hasApiKey", False)),
            },
        },
    }


def find_matching_company_name(requested_name, available_names):
    requested_normalized = normalize_company_name(requested_name)
    for available_name in available_names:
        if normalize_company_name(available_name) == requested_normalized:
            return available_name
    return None


def compute_theme_score(bucket, overall_score):
    bucket_key = (bucket or "X").strip().upper()
    bucket_base = BUCKET_BASE.get(bucket_key, 0)

    try:
        overall_value = float(overall_score or 0)
    except (TypeError, ValueError):
        overall_value = 0.0

    final_score = max(0.0, min(100.0, bucket_base + overall_value))
    return {
        "bucket": bucket_key,
        "overallScore": round(overall_value, 2),
        "finalScore": round(final_score, 2),
    }


def is_not_disclosed_value(value):
    normalized_value = normalize_text_token(str(value or ""))
    return normalized_value in {"", "notdisclosed", "nadisclosed", "na"}


def normalize_commitment_status(status_value):
    normalized_key = normalize_text_token(status_value)
    if normalized_key == "achieved":
        return "achieved", "Achieved"
    if normalized_key == "ontrack":
        return "onTrack", "On-Track"
    if normalized_key == "offtrack":
        return "offTrack", "Off-Track"
    if normalized_key in {"notreporting", "noreporting"}:
        return "noReporting", "Not Reporting"
    return "noTarget", "No-target"


def compute_theme_commitment_progress(theme_data):
    commitments = theme_data.get("Commitments") if isinstance(theme_data, dict) else []
    if not isinstance(commitments, list):
        return {
            "total": 0,
            "achievedOnTrack": 0,
            "pct": 0.0,
        }

    total_commitments = 0
    achieved_on_track = 0

    for commitment_item in commitments:
        if not isinstance(commitment_item, dict):
            continue
        commitment_name = str(commitment_item.get("Name") or "").strip()
        if not commitment_name:
            continue
        total_commitments += 1
        status_key, _ = normalize_commitment_status(commitment_item.get("Status"))
        if status_key in {"achieved", "onTrack"}:
            achieved_on_track += 1

    progress_pct = round((achieved_on_track / total_commitments) * 100, 2) if total_commitments else 0.0
    return {
        "total": total_commitments,
        "achievedOnTrack": achieved_on_track,
        "pct": progress_pct,
    }


def build_best_practice_lookup(sector_name):
    best_practices_dataset = load_scorecard_best_practices_dataset()
    selected_sector_item = next(
        (
            item for item in best_practices_dataset
            if normalize_text_token(item.get("Sector")) == normalize_text_token(sector_name)
        ),
        None,
    )

    if not isinstance(selected_sector_item, dict):
        return {}

    lookup = {}
    for theme_item in selected_sector_item.get("Themes") or []:
        if not isinstance(theme_item, dict):
            continue
        theme_name = str(theme_item.get("Theme") or "").strip()
        if not theme_name:
            continue
        lookup[normalize_text_token(theme_name)] = {
            "bestPlayer": str(theme_item.get("Company") or "").strip(),
            "bestPractice": str(theme_item.get("Practice") or "").strip(),
        }

    return lookup


def resolve_sector_and_company_or_error(user, sector_name, selected_company):
    if not sector_name or not selected_company:
        return None, None, (jsonify({"message": "Sector and company are required."}), 400)

    available_sectors = get_user_available_sectors(user)
    available_sector_names = set(available_sectors)
    if sector_name not in available_sector_names:
        return None, None, (jsonify({"message": "You do not have access to this sector."}), 403)

    sector_companies = SECTOR_COMPANY_MAP.get(sector_name, [])
    matched_selected_company = find_matching_company_name(selected_company, sector_companies)
    if not matched_selected_company:
        return None, None, (jsonify({"message": "Selected company is invalid for this sector."}), 400)

    return sector_companies, matched_selected_company, None


def build_scorecard_context(sector_name, selected_company, sector_companies):
    dataset = load_scorecard_dataset()
    best_practice_lookup = build_best_practice_lookup(sector_name)
    matched_selected_company = find_matching_company_name(selected_company, list(dataset.keys()))
    if not matched_selected_company:
        return None

    selected_company_data = dataset.get(matched_selected_company, {})
    if not isinstance(selected_company_data, dict) or not selected_company_data:
        return None

    peer_company_names = []
    for company_name in sector_companies:
        if normalize_company_name(company_name) == normalize_company_name(selected_company):
            continue
        matched_peer = find_matching_company_name(company_name, list(dataset.keys()))
        if matched_peer:
            peer_company_names.append(matched_peer)

    theme_names = sorted(selected_company_data.keys())
    rows = []
    radar_themes = []
    lagging_candidates = []

    for theme_name in theme_names:
        selected_theme_data = selected_company_data.get(theme_name, {})
        selected_bucket = str(selected_theme_data.get("bucket", "X") or "X").strip().upper()
        if selected_bucket == "X":
            continue

        selected_score = compute_theme_score(
            selected_bucket,
            selected_theme_data.get("overall", 0),
        )

        selected_rationale = selected_theme_data.get("rationale", {})
        rationale_points = []
        if isinstance(selected_rationale, dict):
            rationale_points = [
                str(value).strip()
                for value in selected_rationale.values()
                if str(value).strip()
            ]

        commitments = []
        for commitment_item in selected_theme_data.get("Commitments") or []:
            if not isinstance(commitment_item, dict):
                continue

            commitment_name = str(commitment_item.get("Name") or "").strip()
            commitment_status = str(commitment_item.get("Status") or "").strip()
            commitment_description = commitment_item.get("Description") or {}

            description_points = []
            if isinstance(commitment_description, dict):
                for key_name in ("description", "coverage", "validation"):
                    detail_value = str(commitment_description.get(key_name) or "").strip()
                    if detail_value and not is_not_disclosed_value(detail_value):
                        description_points.append(detail_value)

            if commitment_name:
                commitments.append(
                    {
                        "name": commitment_name,
                        "status": commitment_status,
                        "descriptionPoints": description_points,
                    }
                )

        peer_scores = []
        best_score_value = selected_score["finalScore"]
        best_player = matched_selected_company
        best_practice_source = selected_rationale
        peer_theme_progress_percentages = []
        commitment_peer_status_counts = {}

        for peer_company in peer_company_names:
            peer_theme_data = dataset.get(peer_company, {}).get(theme_name)
            if not isinstance(peer_theme_data, dict):
                continue

            peer_theme_score = compute_theme_score(
                peer_theme_data.get("bucket", "X"),
                peer_theme_data.get("overall", 0),
            )
            peer_scores.append(peer_theme_score["finalScore"])

            if peer_theme_score["finalScore"] > best_score_value:
                best_score_value = peer_theme_score["finalScore"]
                best_player = peer_company
                best_practice_source = peer_theme_data.get("rationale", {})

            peer_progress = compute_theme_commitment_progress(peer_theme_data)
            if peer_progress["total"] > 0:
                peer_theme_progress_percentages.append(peer_progress["pct"])

            for peer_commitment in peer_theme_data.get("Commitments") or []:
                if not isinstance(peer_commitment, dict):
                    continue
                commitment_name = str(peer_commitment.get("Name") or "").strip()
                if not commitment_name:
                    continue

                commitment_key = normalize_text_token(commitment_name)
                if commitment_key not in commitment_peer_status_counts:
                    commitment_peer_status_counts[commitment_key] = {
                        "achieved": 0,
                        "onTrack": 0,
                        "offTrack": 0,
                        "noReporting": 0,
                        "noTarget": 0,
                    }

                status_key, _ = normalize_commitment_status(peer_commitment.get("Status"))
                commitment_peer_status_counts[commitment_key][status_key] += 1

        peer_average = round((sum(peer_scores) / len(peer_scores)), 2) if peer_scores else selected_score["finalScore"]

        gap_vs_peer = round(selected_score["finalScore"] - peer_average, 2)
        lagging_candidates.append({"theme": theme_name, "gap": gap_vs_peer})

        selected_progress = compute_theme_commitment_progress(selected_theme_data)
        company_progress_pct = selected_progress["pct"]
        peer_progress_avg_pct = round(
            (sum(peer_theme_progress_percentages) / len(peer_theme_progress_percentages)),
            2,
        ) if peer_theme_progress_percentages else company_progress_pct

        if abs(company_progress_pct - peer_progress_avg_pct) < 0.01:
            progress_label = "At-Par"
        elif company_progress_pct > peer_progress_avg_pct:
            progress_label = "Leading"
        else:
            progress_label = "Lagging"

        theme_best_practice = best_practice_lookup.get(normalize_text_token(theme_name), {})
        row_best_player = str(theme_best_practice.get("bestPlayer") or "").strip() or best_player
        row_best_practice = str(theme_best_practice.get("bestPractice") or "").strip()
        if not row_best_practice:
            row_best_practice = summarize_best_practice_from_rationale(best_player, best_practice_source, theme_name)

        commitments_with_peer_data = []
        for commitment_item in commitments:
            commitment_name = commitment_item.get("name", "")
            commitment_key = normalize_text_token(commitment_name)
            peer_counts = commitment_peer_status_counts.get(
                commitment_key,
                {
                    "achieved": 0,
                    "onTrack": 0,
                    "offTrack": 0,
                    "noReporting": 0,
                    "noTarget": 0,
                },
            )

            peer_status_parts = []
            if peer_counts["achieved"] > 0:
                peer_status_parts.append(f"{peer_counts['achieved']} achieved")
            if peer_counts["onTrack"] > 0:
                peer_status_parts.append(f"{peer_counts['onTrack']} on track")
            if peer_counts["offTrack"] > 0:
                peer_status_parts.append(f"{peer_counts['offTrack']} off track")
            if peer_counts["noReporting"] > 0:
                peer_status_parts.append(f"{peer_counts['noReporting']} not reporting")
            if peer_counts["noTarget"] > 0:
                peer_status_parts.append(f"{peer_counts['noTarget']} no-target")

            commitments_with_peer_data.append(
                {
                    **commitment_item,
                    "peerStatusSummary": ", ".join(peer_status_parts) if peer_status_parts else "No peer commitments mapped",
                    "peerStatusCounts": peer_counts,
                }
            )

        rows.append(
            {
                "theme": theme_name,
                "overallStatus": selected_score,
                "peerAverage": peer_average,
                "bestScore": round(best_score_value, 2),
                "bestPlayer": row_best_player,
                "bestPractice": row_best_practice,
                "bestPracticeSource": best_practice_source,
                "progress": f"{selected_progress['achievedOnTrack']}/{selected_progress['total']} commitments Achieved or On-Track",
                "progressLabel": progress_label,
                "companyAchievedOnTrackPct": company_progress_pct,
                "peerAchievedOnTrackPct": peer_progress_avg_pct,
                "rationalePoints": rationale_points,
                "commitments": commitments_with_peer_data,
                "commitmentCount": selected_progress["total"],
            }
        )

        radar_themes.append(
            {
                "theme": theme_name,
                "companyScore": selected_score["finalScore"],
                "peerAvg": peer_average,
            }
        )

    sorted_by_gap_ascending = sorted(lagging_candidates, key=lambda item: item["gap"])
    sorted_by_gap_descending = sorted(lagging_candidates, key=lambda item: item["gap"], reverse=True)

    rows.sort(key=lambda item: item.get("overallStatus", {}).get("finalScore", 0), reverse=True)

    lagging_themes = [item["theme"] for item in sorted_by_gap_ascending[:3]]
    leading_themes = []
    for item in sorted_by_gap_descending:
        theme_name = item["theme"]
        if theme_name in lagging_themes:
            continue
        leading_themes.append(theme_name)
        if len(leading_themes) == 3:
            break

    return {
        "selectedCompany": matched_selected_company,
        "peerCompanies": peer_company_names,
        "rows": rows,
        "radarThemes": radar_themes,
        "leadingThemes": leading_themes,
        "laggingThemes": lagging_themes,
        "bucketLegend": load_commitment_scale_dataset().get("bucketLegend", []),
        "industryScaleExample": get_industry_scale_example(sector_name),
    }


def build_fallback_priority_moves(selected_company, lagging_themes):
    if not lagging_themes:
        return [
            f"Prioritize a focused sustainability value-creation agenda for {selected_company} with clear ownership and quarterly milestones.",
            "Tie top initiatives to measurable financial and commercial outcomes to strengthen partner conversations.",
            "Increase transparency on delivery trajectory with externally-facing milestone updates.",
        ]

    moves = [
        f"Launch a 12-month gap-close sprint in {lagging_themes[0]} with explicit KPI targets and operating owners.",
        f"Translate peer-leading practices from {lagging_themes[min(1, len(lagging_themes)-1)]} into a practical execution roadmap.",
        "Build a board-level narrative linking commitment execution to growth, risk resilience, and valuation upside.",
    ]
    return moves


def summarize_best_practice_from_rationale(best_player, rationale, theme_name):
    if isinstance(rationale, dict) and rationale:
        rationale_snippets = [str(value).strip() for value in rationale.values() if str(value).strip()]
        if rationale_snippets:
            return f"{best_player}: {rationale_snippets[0]}"
    return f"{best_player}: Demonstrates comparatively stronger execution signals in {theme_name}."


def generate_openai_insights(selected_company, sector_name, rows, lagging_themes):
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    model_name = os.getenv("OPENAI_MODEL", "gpt-4o-mini").strip() or "gpt-4o-mini"

    default_best_practices = [
        {
            "theme": row["theme"],
            "bestPlayer": row["bestPlayer"],
            "bestPractice": summarize_best_practice_from_rationale(row["bestPlayer"], row.get("bestPracticeSource"), row["theme"]),
        }
        for row in rows
    ]

    if not api_key:
        return {
            "priorityMoves": build_fallback_priority_moves(selected_company, lagging_themes),
            "bestPractices": default_best_practices,
        }

    try:
        from openai import OpenAI

        client = OpenAI(api_key=api_key)

        compact_rows = []
        for row in rows:
            compact_rows.append(
                {
                    "theme": row["theme"],
                    "selectedScore": row["overallStatus"].get("finalScore"),
                    "peerAverage": row["peerAverage"],
                    "bestPlayer": row["bestPlayer"],
                    "bestScore": row["bestScore"],
                    "bestPracticeSource": row.get("bestPracticeSource", {}),
                }
            )

        prompt_payload = {
            "sector": sector_name,
            "selectedCompany": selected_company,
            "rows": compact_rows,
            "instruction": "Return JSON only with keys: priorityMoves (3 concise strings), bestPractices (array of objects with theme, bestPlayer, bestPractice).",
        }

        response = client.chat.completions.create(
            model=model_name,
            temperature=0.2,
            response_format={"type": "json_object"},
            messages=[
                {
                    "role": "system",
                    "content": "You are an expert sustainability strategy advisor for consulting partners. Keep outputs concise, practical, and deal-oriented.",
                },
                {
                    "role": "user",
                    "content": json.dumps(prompt_payload, ensure_ascii=False),
                },
            ],
        )

        content = response.choices[0].message.content or "{}"
        parsed = json.loads(content)

        priority_moves = [str(item).strip() for item in (parsed.get("priorityMoves") or []) if str(item).strip()]
        if len(priority_moves) < 2:
            priority_moves = build_fallback_priority_moves(selected_company, lagging_themes)

        best_practices_raw = parsed.get("bestPractices") or []
        best_practices = []
        for item in best_practices_raw:
            if not isinstance(item, dict):
                continue
            theme = str(item.get("theme") or "").strip()
            best_player = str(item.get("bestPlayer") or "").strip()
            best_practice = str(item.get("bestPractice") or "").strip()
            if theme and best_player and best_practice:
                best_practices.append(
                    {
                        "theme": theme,
                        "bestPlayer": best_player,
                        "bestPractice": best_practice,
                    }
                )

        if not best_practices:
            best_practices = default_best_practices

        return {
            "priorityMoves": priority_moves,
            "bestPractices": best_practices,
        }
    except Exception:
        return {
            "priorityMoves": build_fallback_priority_moves(selected_company, lagging_themes),
            "bestPractices": default_best_practices,
        }


def build_commitment_summary(sector_name, selected_company, sector_companies, scorecard_context):
    breakdown_dataset = load_commitment_breakdown_dataset()
    sector_rows = [
        row for row in breakdown_dataset
        if (row.get("Sector") or "").strip() == sector_name
    ]

    def compute_row_metrics(company_name):
        def read_breakdown_metric(breakdown_obj, aliases):
            if not isinstance(breakdown_obj, dict):
                return 0.0

            normalized_lookup = {}
            for key, value in breakdown_obj.items():
                normalized_key = "".join(character.lower() for character in str(key or "") if character.isalnum())
                normalized_lookup[normalized_key] = value

            for alias in aliases:
                normalized_alias = "".join(character.lower() for character in str(alias or "") if character.isalnum())
                if normalized_alias in normalized_lookup:
                    try:
                        return float(normalized_lookup[normalized_alias] or 0)
                    except (TypeError, ValueError):
                        return 0.0
            return 0.0

        matched = next(
            (
                row for row in sector_rows
                if normalize_company_name(row.get("Company") or "") == normalize_company_name(company_name)
            ),
            None,
        )
        breakdown = (matched or {}).get("Performance Breakdown") or {}
        achieved = read_breakdown_metric(breakdown, ["Achieved"])
        on_track = read_breakdown_metric(breakdown, ["On track", "On-Track", "OnTrack"])
        off_track = read_breakdown_metric(breakdown, ["Off track", "Off-Track", "OffTrack"])
        no_reporting = read_breakdown_metric(breakdown, ["Not Reporting", "No Reporting", "No-reporting", "NoReporting"])
        no_target = read_breakdown_metric(breakdown, ["No-target", "No target", "NoTarget", "Others"])
        total = achieved + on_track + off_track + no_reporting + no_target
        achieved_on_track = achieved + on_track

        achieved_on_track_pct = round((achieved_on_track / total) * 100, 2) if total else 0.0
        off_track_pct = round((off_track / total) * 100, 2) if total else 0.0
        return {
            "company": company_name,
            "isSelected": normalize_company_name(company_name) == normalize_company_name(selected_company),
            "totalCommitments": int(total),
            "achievedOnTrackPct": achieved_on_track_pct,
            "offTrackPct": off_track_pct,
            "breakdown": {
                "achieved": achieved,
                "onTrack": on_track,
                "offTrack": off_track,
                "noReporting": no_reporting,
                "noTarget": no_target,
                "others": no_target,
            },
        }

    selected_metrics = compute_row_metrics(selected_company)
    peer_metrics = [
        compute_row_metrics(company_name)
        for company_name in sector_companies
        if normalize_company_name(company_name) != normalize_company_name(selected_company)
    ]
    peer_metrics.sort(key=lambda item: item["achievedOnTrackPct"], reverse=True)

    ranking = [selected_metrics, *peer_metrics]

    return {
        "sector": sector_name,
        "selectedCompany": selected_company,
        "totalCommitments": selected_metrics["totalCommitments"],
        "achievedOnTrackPct": selected_metrics["achievedOnTrackPct"],
        "offTrackPct": selected_metrics["offTrackPct"],
        "leadingThemes": scorecard_context.get("leadingThemes", []),
        "laggingThemes": scorecard_context.get("laggingThemes", []),
        "ranking": ranking,
        "themeScores": scorecard_context.get("radarThemes", []),
    }


@app.route("/api/auth/login", methods=["POST"])
def login():
    payload = request.get_json(silent=True) or {}
    email = (payload.get("email") or "").strip().lower()
    password = payload.get("password") or ""

    if not email or not password:
        return jsonify({"message": "Email and password are required."}), 400

    user = get_user_by_email(email)
    if not user or not check_password_hash(user.password_hash, password):
        return jsonify({"message": "Invalid credentials."}), 401

    session["user_id"] = user.id

    return jsonify({
        "message": "Login successful.",
        "user": user.to_dict(),
    })


@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"message": "Logged out."})


@app.route("/api/auth/me", methods=["GET"])
def me():
    user = get_current_user()
    if not user:
        return jsonify({"message": "Unauthorized."}), 401

    return jsonify({"user": user.to_dict()})


@app.route("/api/auth/change-password", methods=["POST"])
def change_password():
    user = get_current_user()
    if not user:
        return jsonify({"message": "Unauthorized."}), 401

    payload = request.get_json(silent=True) or {}
    current_password = payload.get("currentPassword") or ""
    new_password = payload.get("newPassword") or ""

    if not current_password or not new_password:
        return jsonify({"message": "Current and new passwords are required."}), 400

    if len(new_password) < 6:
        return jsonify({"message": "New password must be at least 6 characters."}), 400

    if not check_password_hash(user.password_hash, current_password):
        return jsonify({"message": "Current password is incorrect."}), 400

    user.password_hash = generate_password_hash(new_password)
    save_user(user)

    return jsonify({"message": "Password changed successfully.", "user": user.to_dict()})


@app.route("/api/profile", methods=["GET", "PUT"])
def profile():
    user = get_current_user()
    if not user:
        return jsonify({"message": "Unauthorized."}), 401

    if request.method == "GET":
        return jsonify({"user": user.to_dict()})

    payload = request.get_json(silent=True) or {}
    first_name = (payload.get("firstName") or "").strip()
    last_name = (payload.get("lastName") or "").strip()

    if not first_name or not last_name:
        return jsonify({"message": "First name and last name are required."}), 400

    user.first_name = first_name
    user.last_name = last_name
    save_user(user)

    return jsonify({"message": "Profile updated successfully.", "user": user.to_dict()})


@app.route("/api/options", methods=["GET"])
def options():
    user = get_current_user()
    if not user:
        return jsonify({"message": "Unauthorized."}), 401

    available_sectors = get_user_available_sectors(user)
    sector_names = list(available_sectors)
    sector_company_map = {
        sector_name: get_ordered_company_names(sector_name)
        for sector_name in available_sectors
    }

    return jsonify({"sectors": sector_names, "sectorCompanyMap": sector_company_map})


@app.route("/api/material-topics/comparison", methods=["GET"])
def material_topics_comparison():
    user = get_current_user()
    if not user:
        return jsonify({"message": "Unauthorized."}), 401

    sector_name = (request.args.get("sector") or "").strip()
    selected_company = (request.args.get("company") or "").strip()

    if not sector_name or not selected_company:
        return jsonify({"message": "Sector and company are required."}), 400

    available_sector_names = set(get_user_available_sectors(user))
    if sector_name not in available_sector_names:
        return jsonify({"message": "You do not have access to this sector."}), 403

    dataset = load_material_topics_dataset()
    sector_rows = [row for row in dataset if (row.get("Sector") or "").strip() == sector_name]
    if not sector_rows:
        return jsonify({"message": "No material topic data found for the selected sector."}), 404

    selected_row = next(
        (row for row in sector_rows if (row.get("Company") or "").strip() == selected_company),
        None,
    )
    if not selected_row:
        return jsonify({"message": "No material topic data found for the selected company."}), 404

    competitor_rows = [row for row in sector_rows if (row.get("Company") or "").strip() != selected_company]
    competitor_companies = [(row.get("Company") or "").strip() for row in competitor_rows]

    competitor_topic_sets = {
        company_name: {
            (topic or "").strip()
            for topic in ((row.get("Material Topics") or []))
            if (topic or "").strip()
        }
        for company_name, row in zip(competitor_companies, competitor_rows)
    }

    selected_topics = [
        (topic or "").strip()
        for topic in (selected_row.get("Material Topics") or [])
        if (topic or "").strip()
    ]

    rows = []
    for topic in selected_topics:
        competitor_matches = {
            company_name: topic in competitor_topic_sets.get(company_name, set())
            for company_name in competitor_companies
        }
        match_count = sum(1 for value in competitor_matches.values() if value)
        rows.append(
            {
                "materialTopic": topic,
                "matchCount": match_count,
                "competitorMatches": competitor_matches,
            }
        )

    rows.sort(key=lambda row: (-row["matchCount"], row["materialTopic"].lower()))

    return jsonify(
        {
            "sector": sector_name,
            "selectedCompany": selected_company,
            "competitorCompanies": competitor_companies,
            "rows": rows,
        }
    )


@app.route("/api/recommendations", methods=["GET"])
def recommendations():
    user = get_current_user()
    if not user:
        return jsonify({"message": "Unauthorized."}), 401

    sector_name = (request.args.get("sector") or "").strip()
    selected_company = (request.args.get("company") or "").strip()

    if not sector_name or not selected_company:
        return jsonify({"message": "Sector and company are required."}), 400

    available_sector_names = set(get_user_available_sectors(user))
    if sector_name not in available_sector_names:
        return jsonify({"message": "You do not have access to this sector."}), 403

    dataset = load_recommendations_dataset()
    selected_row = next(
        (
            row for row in dataset
            if (row.get("Sector") or "").strip() == sector_name
            and (row.get("Company") or "").strip() == selected_company
        ),
        None,
    )

    if not selected_row:
        return jsonify({"message": "No recommendations found for this company."}), 404

    actions = []
    for action in selected_row.get("Strategic Actions") or []:
        topic = (action.get("Topic") or "").strip()
        details = [
            (detail or "").strip()
            for detail in (action.get("Details") or [])
            if (detail or "").strip()
        ]
        if topic:
            actions.append({"topic": topic, "details": details})

    return jsonify({
        "sector": sector_name,
        "selectedCompany": selected_company,
        "actions": actions,
    })


@app.route("/api/commitments/overview", methods=["GET"])
def commitments_overview():
    user = get_current_user()
    if not user:
        return jsonify({"message": "Unauthorized."}), 401

    sector_name = (request.args.get("sector") or "").strip()
    selected_company = (request.args.get("company") or "").strip()

    sector_companies, matched_company, error = resolve_sector_and_company_or_error(user, sector_name, selected_company)
    if error:
        return error

    scorecard_context = build_scorecard_context(sector_name, matched_company, sector_companies)
    if not scorecard_context:
        return jsonify({"message": "Scorecard data not available for selected company."}), 404

    summary = build_commitment_summary(sector_name, matched_company, sector_companies, scorecard_context)
    return jsonify(summary)


@app.route("/api/scorecard", methods=["GET"])
def scorecard():
    user = get_current_user()
    if not user:
        return jsonify({"message": "Unauthorized."}), 401

    sector_name = (request.args.get("sector") or "").strip()
    selected_company = (request.args.get("company") or "").strip()

    sector_companies, matched_company, error = resolve_sector_and_company_or_error(user, sector_name, selected_company)
    if error:
        return error

    scorecard_context = build_scorecard_context(sector_name, matched_company, sector_companies)
    if not scorecard_context:
        return jsonify({"message": "Scorecard data not available for selected company."}), 404

    ai_insights = generate_openai_insights(
        matched_company,
        sector_name,
        scorecard_context["rows"],
        scorecard_context.get("laggingThemes", []),
    )

    rows = []
    for row in scorecard_context["rows"]:
        rows.append(
            {
                "theme": row["theme"],
                "overallStatus": row["overallStatus"],
                "peerAverage": row["peerAverage"],
                "bestScore": row["bestScore"],
                "bestPlayer": row["bestPlayer"],
                "bestPractice": row["bestPractice"],
                "progress": row["progress"],
                "progressLabel": row.get("progressLabel"),
                "companyAchievedOnTrackPct": row.get("companyAchievedOnTrackPct", 0),
                "peerAchievedOnTrackPct": row.get("peerAchievedOnTrackPct", 0),
                "rationalePoints": row.get("rationalePoints", []),
                "commitments": row.get("commitments", []),
                "commitmentCount": row.get("commitmentCount", 0),
                "themeScale": get_theme_scale(sector_name, row["theme"]),
            }
        )

    return jsonify(
        {
            "sector": sector_name,
            "selectedCompany": matched_company,
            "priorityMoves": ai_insights.get("priorityMoves", []),
            "radarThemes": scorecard_context["radarThemes"],
            "bucketLegend": scorecard_context.get("bucketLegend", []),
            "industryScaleExample": scorecard_context.get("industryScaleExample"),
            "rows": rows,
        }
    )


@app.route("/api/investments/insights", methods=["GET"])
def investment_insights():
    user = get_current_user()
    if not user:
        return jsonify({"message": "Unauthorized."}), 401

    sector_name = (request.args.get("sector") or "").strip()
    selected_company = (request.args.get("company") or "").strip()

    sector_companies, matched_company, error = resolve_sector_and_company_or_error(user, sector_name, selected_company)
    if error:
        return error

    insights = build_investment_insights(sector_name, matched_company, sector_companies)
    return jsonify(insights)


@app.route("/api/version", methods=["GET"])
def app_version():
    return jsonify({"version": APP_VERSION})


@app.route("/api/admin/metadata", methods=["GET"])
def admin_metadata():
    _, error = get_admin_user_or_error()
    if error:
        return error

    all_sectors = list(SECTOR_COMPANY_MAP.keys())
    return jsonify({
        "sectors": all_sectors,
        "userStore": {
            "path": USERS_DATA_PATH,
            "ephemeral": is_ephemeral_user_store(),
        },
    })


@app.route("/api/admin/users", methods=["GET", "POST"])
def admin_users():
    _, error = get_admin_user_or_error()
    if error:
        return error

    if request.method == "GET":
        users = list_users()
        return jsonify({"users": [user.to_admin_dict() for user in users]})

    payload = request.get_json(silent=True) or {}
    email = (payload.get("email") or "").strip().lower()
    password = payload.get("password") or ""
    first_name = (payload.get("firstName") or "").strip()
    last_name = (payload.get("lastName") or "").strip()
    sector_names = payload.get("sectors") or []
    is_admin = bool(payload.get("isAdmin", False))

    if not email or not password or not first_name or not last_name:
        return jsonify({"message": "Email, password, first name and last name are required."}), 400

    if len(password) < 6:
        return jsonify({"message": "Password must be at least 6 characters."}), 400

    if get_user_by_email(email):
        return jsonify({"message": "User with this email already exists."}), 400

    sectors, missing_sectors = resolve_sector_names(sector_names)
    if missing_sectors:
        return jsonify({"message": f"Invalid sectors: {', '.join(missing_sectors)}"}), 400

    new_user = create_user_record(
        email=email,
        password=password,
        first_name=first_name,
        last_name=last_name,
        sector_names=sectors,
        is_admin=is_admin,
    )
    if not new_user:
        return jsonify({"message": "User with this email already exists."}), 400

    response = {
        "message": "User created successfully.",
        "user": new_user.to_admin_dict(),
    }
    if is_ephemeral_user_store():
        response["storageWarning"] = "User data is stored in ephemeral serverless storage and may not persist across instances/deployments."

    return jsonify(response)


@app.route("/api/admin/users/<int:user_id>", methods=["PUT", "DELETE"])
def admin_user_by_id(user_id):
    admin_user, error = get_admin_user_or_error()
    if error:
        return error
    if not admin_user:
        return jsonify({"message": "Unauthorized."}), 401

    target_user = get_user_by_id(user_id)
    if not target_user:
        return jsonify({"message": "User not found."}), 404

    if request.method == "DELETE":
        if target_user.id == admin_user.id:
            return jsonify({"message": "You cannot delete your own account."}), 400
        delete_user(target_user.id)
        response = {"message": "User deleted successfully."}
        if is_ephemeral_user_store():
            response["storageWarning"] = "User data is stored in ephemeral serverless storage and may not persist across instances/deployments."
        return jsonify(response)

    payload = request.get_json(silent=True) or {}
    first_name = (payload.get("firstName") or "").strip()
    last_name = (payload.get("lastName") or "").strip()
    password = (payload.get("password") or "").strip()
    sector_names = payload.get("sectors")
    is_admin = payload.get("isAdmin")

    if not first_name or not last_name:
        return jsonify({"message": "First name and last name are required."}), 400

    target_user.first_name = first_name
    target_user.last_name = last_name

    if isinstance(is_admin, bool):
        if target_user.id == admin_user.id and not is_admin:
            return jsonify({"message": "You cannot remove your own admin access."}), 400
        target_user.is_admin = is_admin

    if password:
        if len(password) < 6:
            return jsonify({"message": "Password must be at least 6 characters."}), 400
        target_user.password_hash = generate_password_hash(password)

    if sector_names is not None:
        sectors, missing_sectors = resolve_sector_names(sector_names)
        if missing_sectors:
            return jsonify({"message": f"Invalid sectors: {', '.join(missing_sectors)}"}), 400
        target_user.sectors = sectors

    save_user(target_user)

    response = {"message": "User updated successfully.", "user": target_user.to_admin_dict()}
    if is_ephemeral_user_store():
        response["storageWarning"] = "User data is stored in ephemeral serverless storage and may not persist across instances/deployments."

    return jsonify(response)


ensure_user_store()


if __name__ == "__main__":
    app.run(debug=True)
